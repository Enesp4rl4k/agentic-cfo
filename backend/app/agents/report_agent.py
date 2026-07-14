"""
Report Agent — Skill 5 of 8 (final output node).

Responsibility: Generate Excel report and dashboard JSON from all agent outputs:
P&L, Cash Flow, Forecast, Tax Analysis, Anomaly Detection, Budget Comparison.

done_when: state['report_paths']['xlsx'] exists on disk AND state['dashboard_json'] is populated.
"""
from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from typing import Any

from app.agents.state import CFOState, AgentRunConfig, SkillResult
from app.config import get_settings

logger = logging.getLogger(__name__)


def _fmt(cents: int | float) -> float:
    return round(cents / 100, 2)


def _build_dashboard_json(state: CFOState) -> dict[str, Any]:
    """Build the JSON payload consumed by the frontend dashboard."""
    pnl = state.get("pnl", {})
    cashflow = state.get("cashflow", {})
    forecast = state.get("forecast", {})
    tax_analysis = state.get("tax_analysis", {})
    anomalies = state.get("anomalies", {})
    budget_comparison = state.get("budget_comparison", {})
    transactions = state.get("transactions", [])

    # KPI cards
    kpis = [
        {"label": "Revenue", "value": _fmt(pnl.get("revenue", 0)), "format": "currency", "trend": None},
        {"label": "Net Income", "value": _fmt(pnl.get("net_income", 0)), "format": "currency", "trend": None},
        {"label": "Gross Margin", "value": round(pnl.get("gross_margin", 0) * 100, 1), "format": "percent", "trend": None},
        {"label": "Net Cash Flow", "value": _fmt(cashflow.get("net_change", 0)), "format": "currency", "trend": None},
        {"label": "EBITDA", "value": _fmt(pnl.get("ebitda", 0)), "format": "currency", "trend": None},
    ]

    # Runway KPI
    base_scenario = forecast.get("scenarios", {}).get("base", {})
    if base_scenario.get("runway_months") is not None:
        kpis.append({"label": "Cash Runway", "value": base_scenario["runway_months"], "format": "months", "trend": None})

    # Tax KPI
    if tax_analysis:
        kpis.append({
            "label": "Tax Burden",
            "value": _fmt(tax_analysis.get("total_tax_burden", 0)),
            "format": "currency",
            "trend": None,
        })

    # Anomaly risk KPI
    if anomalies:
        kpis.append({
            "label": "Risk Score",
            "value": round(anomalies.get("risk_score", 0) * 100, 0),
            "format": "score",
            "trend": None,
        })

    # Recent transactions (last 20)
    recent_transactions = sorted(
        transactions,
        key=lambda t: t.get("transaction_date") or "",
        reverse=True,
    )[:20]

    # All alerts combined
    all_alerts = (
        list(cashflow.get("alerts", []))
        + list(forecast.get("alerts", []))
        + list(tax_analysis.get("alerts", []) if tax_analysis else [])
        + list(budget_comparison.get("alerts", []) if budget_comparison else [])
    )

    # Anomaly alerts (top 3 high severity)
    if anomalies:
        for a in anomalies.get("anomaly_list", []):
            if a.get("severity") == "high":
                all_alerts.append({
                    "level": "critical",
                    "message": f"[Anomali] {a.get('detail', '')}",
                })

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "kpis": kpis,
        "pnl": {
            "revenue": _fmt(pnl.get("revenue", 0)),
            "cogs": _fmt(pnl.get("cogs", 0)),
            "gross_profit": _fmt(pnl.get("gross_profit", 0)),
            "gross_margin": pnl.get("gross_margin", 0),
            "ebitda": _fmt(pnl.get("ebitda", 0)),
            "ebitda_margin": pnl.get("ebitda_margin", 0),
            "net_income": _fmt(pnl.get("net_income", 0)),
            "net_margin": pnl.get("net_margin", 0),
            "opex": {k: _fmt(v) for k, v in pnl.get("opex", {}).items()},
            "narrative": pnl.get("narrative", ""),
        },
        "cashflow": {
            "operating": _fmt(cashflow.get("operating", 0)),
            "investing": _fmt(cashflow.get("investing", 0)),
            "financing": _fmt(cashflow.get("financing", 0)),
            "net_change": _fmt(cashflow.get("net_change", 0)),
            "monthly_series": cashflow.get("monthly_series", []),
            "narrative": cashflow.get("narrative", ""),
            "alerts": cashflow.get("alerts", []),
        },
        "forecast": {
            "scenarios": {
                k: {
                    "label": v.get("label"),
                    "description": v.get("description"),
                    "runway_months": v.get("runway_months"),
                    "twelve_month_net": _fmt(v.get("twelve_month_net", 0)),
                    "months": v.get("months", []),
                }
                for k, v in forecast.get("scenarios", {}).items()
            },
            "narrative": forecast.get("narrative", ""),
            "alerts": forecast.get("alerts", []),
        },
        "tax_analysis": {
            "kdv_collected": _fmt(tax_analysis.get("kdv", {}).get("kdv_collected", 0)),
            "kdv_paid": _fmt(tax_analysis.get("kdv", {}).get("kdv_paid", 0)),
            "kdv_net": _fmt(tax_analysis.get("kdv", {}).get("kdv_net", 0)),
            "kdv_payable": _fmt(tax_analysis.get("kdv", {}).get("kdv_payable", 0)),
            "kdv_refundable": _fmt(tax_analysis.get("kdv", {}).get("kdv_refundable", 0)),
            "monthly_kdv": tax_analysis.get("kdv", {}).get("monthly_kdv", []),
            "stopaj_total": _fmt(tax_analysis.get("stopaj", {}).get("stopaj_total", 0)),
            "stopaj_salary": _fmt(tax_analysis.get("stopaj", {}).get("stopaj_salary", 0)),
            "stopaj_rent": _fmt(tax_analysis.get("stopaj", {}).get("stopaj_rent", 0)),
            "kurumlar_vergisi_annual": _fmt(
                tax_analysis.get("kurumlar_vergisi", {}).get("kurumlar_vergisi_annual", 0)
            ),
            "gecici_vergi_quarterly": _fmt(
                tax_analysis.get("kurumlar_vergisi", {}).get("gecici_vergi_quarterly", 0)
            ),
            "total_tax_burden": _fmt(tax_analysis.get("total_tax_burden", 0)),
            "effective_tax_rate": tax_analysis.get("effective_tax_rate", 0),
            "narrative": tax_analysis.get("narrative", ""),
            "alerts": tax_analysis.get("alerts", []),
        } if tax_analysis else None,
        "anomalies": {
            "anomaly_list": anomalies.get("anomaly_list", []),
            "anomaly_count": anomalies.get("anomaly_count", 0),
            "high_severity_count": anomalies.get("high_severity_count", 0),
            "risk_score": anomalies.get("risk_score", 0.0),
            "narrative": anomalies.get("narrative", ""),
        } if anomalies else None,
        "budget_comparison": {
            "categories": {
                cat: {
                    "budget": _fmt(d.get("budget", 0)),
                    "actual": _fmt(d.get("actual", 0)),
                    "variance": _fmt(d.get("variance", 0)),
                    "variance_pct": d.get("variance_pct", 0),
                    "status": d.get("status"),
                }
                for cat, d in budget_comparison.get("categories", {}).items()
            },
            "total_variance": _fmt(budget_comparison.get("total_variance", 0)),
            "variance_pct": budget_comparison.get("variance_pct", 0),
            "over_budget_count": budget_comparison.get("over_budget_count", 0),
            "auto_budget": budget_comparison.get("auto_budget", True),
            "narrative": budget_comparison.get("narrative", ""),
            "alerts": budget_comparison.get("alerts", []),
        } if budget_comparison else None,
        "alerts": all_alerts,
        "recent_transactions": recent_transactions,
        "transaction_count": len(transactions),
    }


def _write_excel(
    pnl: dict[str, Any],
    cashflow: dict[str, Any],
    forecast: dict[str, Any],
    tax_analysis: dict[str, Any] | None,
    anomalies: dict[str, Any] | None,
    budget_comparison: dict[str, Any] | None,
    output_path: str,
) -> None:
    """Write a multi-sheet Excel report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SUBHEADER_FILL = PatternFill("solid", fgColor="E8F0FE")

    wb = Workbook()

    # ── Sheet 1: P&L ──────────────────────────────────────────────────────────
    ws_pnl = wb.active
    ws_pnl.title = "P&L Statement"
    ws_pnl.append(["Item", "Amount (₺)"])
    for cell in ws_pnl[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    pnl_rows = [
        ("Revenue", pnl.get("revenue", 0)),
        ("Cost of Goods Sold (COGS)", -pnl.get("cogs", 0)),
        ("Gross Profit", pnl.get("gross_profit", 0)),
        ("Gross Margin %", pnl.get("gross_margin", 0) * 100),
        ("", None),
        ("Operating Expenses", None),
        *[(f"  {k.replace('_', ' ').title()}", -v) for k, v in pnl.get("opex", {}).items()],
        ("Total OpEx", -pnl.get("total_opex", 0)),
        ("", None),
        ("EBITDA", pnl.get("ebitda", 0)),
        ("EBITDA Margin %", pnl.get("ebitda_margin", 0) * 100),
        ("Tax", -pnl.get("tax", 0)),
        ("Loan Payments", -pnl.get("loan_payments", 0)),
        ("Net Income", pnl.get("net_income", 0)),
        ("Net Margin %", pnl.get("net_margin", 0) * 100),
    ]

    for label, value in pnl_rows:
        if value is None:
            ws_pnl.append([label, ""])
        elif isinstance(value, float) and "%" in label:
            ws_pnl.append([label, round(value, 2)])
        elif isinstance(value, (int, float)):
            ws_pnl.append([label, _fmt(int(value))])
        else:
            ws_pnl.append([label, value])

    ws_pnl.column_dimensions["A"].width = 35
    ws_pnl.column_dimensions["B"].width = 18

    if pnl.get("narrative"):
        ws_pnl.append([])
        ws_pnl.append(["CFO Commentary"])
        ws_pnl[-1][0].font = Font(bold=True)
        ws_pnl.append([pnl["narrative"]])
        ws_pnl[ws_pnl.max_row][0].alignment = Alignment(wrap_text=True)
        ws_pnl.row_dimensions[ws_pnl.max_row].height = 80

    # ── Sheet 2: Cash Flow ────────────────────────────────────────────────────
    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.append(["Activity", "Amount (₺)"])
    for cell in ws_cf[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    cf_rows = [
        ("Operating Cash Flow", cashflow.get("operating", 0)),
        ("  Cash Inflows", cashflow.get("operating_in", 0)),
        ("  Cash Outflows", -cashflow.get("operating_out", 0)),
        ("Investing Cash Flow", cashflow.get("investing", 0)),
        ("Financing Cash Flow", cashflow.get("financing", 0)),
        ("Net Cash Change", cashflow.get("net_change", 0)),
    ]
    for label, value in cf_rows:
        ws_cf.append([label, _fmt(int(value))])

    ws_cf.append([])
    ws_cf.append(["Monthly Cash Flow"])
    ws_cf[-1][0].font = Font(bold=True)
    ws_cf.append(["Month", "Cash In (₺)", "Cash Out (₺)", "Net (₺)"])
    for cell in ws_cf[ws_cf.max_row]:
        cell.fill = SUBHEADER_FILL
    for entry in cashflow.get("monthly_series", []):
        ws_cf.append([
            entry["month"],
            _fmt(entry["in"]),
            _fmt(entry["out"]),
            _fmt(entry["net"]),
        ])

    ws_cf.column_dimensions["A"].width = 30
    for col in ["B", "C", "D"]:
        ws_cf.column_dimensions[col].width = 16

    # ── Sheet 3: Forecast ─────────────────────────────────────────────────────
    ws_fc = wb.create_sheet("Forecast")
    ws_fc.append(["12-Month Financial Forecast"])
    ws_fc[1][0].font = Font(bold=True, size=13)
    ws_fc.append([])

    for scenario in forecast.get("scenarios", {}).values():
        ws_fc.append([scenario["label"], scenario.get("description", "")])
        ws_fc[-1][0].font = Font(bold=True)
        ws_fc.append(["Month", "Cash In (₺)", "Cash Out (₺)", "Net (₺)"])
        for cell in ws_fc[ws_fc.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for entry in scenario.get("months", []):
            ws_fc.append([entry["month"], _fmt(entry["in"]), _fmt(entry["out"]), _fmt(entry["net"])])
        ws_fc.append(["12-Month Net", "", "", _fmt(scenario.get("twelve_month_net", 0))])
        ws_fc.append(["Cash Runway", f"{scenario.get('runway_months', 'Stable')} months", "", ""])
        ws_fc.append([])

    ws_fc.column_dimensions["A"].width = 20
    for col in ["B", "C", "D"]:
        ws_fc.column_dimensions[col].width = 16

    # ── Sheet 4: Tax Analysis ─────────────────────────────────────────────────
    if tax_analysis:
        ws_tax = wb.create_sheet("Tax Analysis")
        ws_tax.append(["Vergi Analizi", ""])
        ws_tax[1][0].font = Font(bold=True, size=13)
        ws_tax.append([])

        ws_tax.append(["KDV (VAT)", ""])
        ws_tax[-1][0].font = Font(bold=True)
        ws_tax[-1][0].fill = SUBHEADER_FILL
        kdv = tax_analysis.get("kdv", {})
        ws_tax.append(["  Tahsil Edilen KDV", _fmt(kdv.get("kdv_collected", 0))])
        ws_tax.append(["  Ödenen KDV", _fmt(kdv.get("kdv_paid", 0))])
        ws_tax.append(["  KDV Net (Ödenecek)", _fmt(kdv.get("kdv_net", 0))])
        ws_tax.append([])

        ws_tax.append(["Stopaj (Withholding Tax)", ""])
        ws_tax[-1][0].font = Font(bold=True)
        ws_tax[-1][0].fill = SUBHEADER_FILL
        stopaj = tax_analysis.get("stopaj", {})
        ws_tax.append(["  Maaş Stopajı", _fmt(stopaj.get("stopaj_salary", 0))])
        ws_tax.append(["  Kira Stopajı", _fmt(stopaj.get("stopaj_rent", 0))])
        ws_tax.append(["  Toplam Stopaj", _fmt(stopaj.get("stopaj_total", 0))])
        ws_tax.append([])

        ws_tax.append(["Kurumlar Vergisi", ""])
        ws_tax[-1][0].font = Font(bold=True)
        ws_tax[-1][0].fill = SUBHEADER_FILL
        kv = tax_analysis.get("kurumlar_vergisi", {})
        ws_tax.append(["  Vergilendirilebilir Gelir", _fmt(kv.get("taxable_income", 0))])
        ws_tax.append(["  Vergi Oranı (%25)", "25%"])
        ws_tax.append(["  Yıllık Kurumlar Vergisi", _fmt(kv.get("kurumlar_vergisi_annual", 0))])
        ws_tax.append(["  Geçici Vergi (Çeyrek)", _fmt(kv.get("gecici_vergi_quarterly", 0))])
        ws_tax.append([])

        ws_tax.append(["TOPLAM VERGİ YÜKÜ", _fmt(tax_analysis.get("total_tax_burden", 0))])
        ws_tax[-1][0].font = Font(bold=True)

        if tax_analysis.get("narrative"):
            ws_tax.append([])
            ws_tax.append(["Mali Müşavir Yorumu"])
            ws_tax[-1][0].font = Font(bold=True)
            ws_tax.append([tax_analysis["narrative"]])
            ws_tax[ws_tax.max_row][0].alignment = Alignment(wrap_text=True)
            ws_tax.row_dimensions[ws_tax.max_row].height = 80

        ws_tax.column_dimensions["A"].width = 40
        ws_tax.column_dimensions["B"].width = 20

    # ── Sheet 5: Anomalies ────────────────────────────────────────────────────
    if anomalies and anomalies.get("anomaly_list"):
        ws_anom = wb.create_sheet("Anomalies")
        ws_anom.append(["Anomali Tespiti"])
        ws_anom[1][0].font = Font(bold=True, size=13)
        ws_anom.append([f"Risk Skoru: {anomalies.get('risk_score', 0):.0%}  |  Toplam: {anomalies.get('anomaly_count', 0)}"])
        ws_anom.append([])

        ws_anom.append(["Tür", "Önem", "Tarih", "Tutar (₺)", "Tedarikçi", "Açıklama"])
        for cell in ws_anom[ws_anom.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for a in anomalies.get("anomaly_list", []):
            ws_anom.append([
                a.get("type", ""),
                a.get("severity", ""),
                str(a.get("transaction_date", ""))[:10],
                _fmt(a.get("amount_cents", 0)),
                a.get("vendor", ""),
                a.get("detail", ""),
            ])

        ws_anom.column_dimensions["A"].width = 20
        ws_anom.column_dimensions["B"].width = 10
        ws_anom.column_dimensions["C"].width = 12
        ws_anom.column_dimensions["D"].width = 15
        ws_anom.column_dimensions["E"].width = 25
        ws_anom.column_dimensions["F"].width = 60

    # ── Sheet 6: Budget Comparison ────────────────────────────────────────────
    if budget_comparison and budget_comparison.get("categories"):
        ws_bgt = wb.create_sheet("Budget vs Actual")
        ws_bgt.append(["Bütçe-Gerçekleşme Karşılaştırması"])
        ws_bgt[1][0].font = Font(bold=True, size=13)
        auto = budget_comparison.get("auto_budget", True)
        ws_bgt.append([f"Bütçe tipi: {'Otomatik (Gerçekleşmeden)' if auto else 'Manuel'}"])
        ws_bgt.append([])

        ws_bgt.append(["Kategori", "Bütçe (₺)", "Gerçekleşen (₺)", "Fark (₺)", "Fark (%)", "Durum"])
        for cell in ws_bgt[ws_bgt.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for cat, d in budget_comparison.get("categories", {}).items():
            ws_bgt.append([
                cat.replace("_", " ").title(),
                _fmt(d.get("budget", 0)),
                _fmt(d.get("actual", 0)),
                _fmt(d.get("variance", 0)),
                f"{d.get('variance_pct', 0):.1f}%",
                d.get("status", ""),
            ])

        ws_bgt.append([])
        ws_bgt.append(["TOPLAM VARYANS", "", "", _fmt(budget_comparison.get("total_variance", 0)), f"{budget_comparison.get('variance_pct', 0):.1f}%", ""])
        ws_bgt[-1][0].font = Font(bold=True)

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_bgt.column_dimensions[col].width = 22

    wb.save(output_path)


async def run_report(state: CFOState, config: AgentRunConfig) -> SkillResult:
    """
    Report Skill.
    done_when: state['report_paths']['xlsx'] exists AND state['dashboard_json'] is populated.
    """
    pnl = state.get("pnl", {})
    cashflow = state.get("cashflow", {})
    forecast = state.get("forecast", {})

    if not pnl or not cashflow:
        return SkillResult(
            ok=False,
            detail="P&L or cash flow data missing — cannot generate report.",
            halt=True,
        )

    try:
        settings = get_settings()
        job_id = state.get("job_id", "unknown")

        output_dir = os.path.join(settings.storage_local_path, "reports", job_id)
        os.makedirs(output_dir, exist_ok=True)

        xlsx_path = os.path.join(output_dir, "financial_report.xlsx")
        _write_excel(
            pnl=pnl,
            cashflow=cashflow,
            forecast=forecast,
            tax_analysis=state.get("tax_analysis"),
            anomalies=state.get("anomalies"),
            budget_comparison=state.get("budget_comparison"),
            output_path=xlsx_path,
        )

        dashboard_json = _build_dashboard_json(state)

        return SkillResult(
            ok=True,
            patch={
                "report_paths": {"xlsx": xlsx_path},
                "dashboard_json": dashboard_json,
            },
            confidence=1.0,
            detail=f"Report generated: {xlsx_path}",
        )
    except Exception as exc:
        logger.exception("Report agent failed for job=%s", state.get("job_id"))
        return SkillResult(ok=False, detail=f"Report error: {exc}")
