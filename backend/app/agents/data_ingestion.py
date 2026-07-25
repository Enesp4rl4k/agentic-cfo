"""
Data Ingestion Agent — Skill 1 of 5.

Responsibility: Read the uploaded file (PDF/Excel/CSV), extract raw text
and structured transactions, persist them to the DB.

Parse strategy (fast → slow):
  1. Bank-specific parser (ParserRegistry.detect) — rule-based, free, fast
  2. LLM extraction (_extract_transactions_with_llm) — fallback only

Confidence signals:
- 1.0  → all transactions parsed with vendor + amount + date
- 0.85 → some fields missing but majority parseable
- 0.60 → low-quality OCR or too many unparseable rows → triggers review gate
"""
from __future__ import annotations

import logging
import os
import re
from datetime import datetime, timezone
from typing import Any

from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

from app.agents.state import CFOState, AgentRunConfig, SkillResult
from app.config import get_settings
from app.parsers.registry import ParserRegistry
from app.parsers.base import ParsedStatement

logger = logging.getLogger(__name__)

CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "revenue":        ["sales", "income", "revenue", "payment received", "invoice issued"],
    "cogs":           ["raw material", "goods", "product cost", "manufacturing", "inventory"],
    "salary":         ["salary", "payroll", "wages", "employee", "social security", "staff"],
    "rent":           ["rent", "lease", "office rent"],
    "utilities":      ["electricity", "water", "gas", "internet", "phone", "telecom"],
    "marketing":      ["advertising", "marketing", "google ads", "meta ads", "social media", "campaign"],
    "technology":     ["software", "server", "cloud", "aws", "azure", "saas", "subscription", "license"],
    "tax":            ["tax", "vat", "withholding", "corporate tax", "income tax"],
    "loan":           ["loan", "debt", "interest", "installment", "bank payment", "credit"],
    "other_expense":  [],
    "other_income":   [],
}


def _guess_category(description: str) -> str:
    desc_lower = description.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(kw in desc_lower for kw in keywords):
            return category
    return "other_expense"


def _parse_amount(raw: str) -> int | None:
    """Parse an amount string into the smallest currency unit (cents/kurus).

    Handles:
      - "1000.50"  → 100050
      - "1.000,50" → 100050 (European format: dot=thousands, comma=decimal)
      - "1,000.50" → 100050 (Anglo format: comma=thousands, dot=decimal)
      - "₺1,500"   → 150000 (TRY symbol + Anglo format)
    """
    # Strip currency symbols and whitespace
    s = re.sub(r"[^\d,.]", "", raw)
    if not s:
        return None

    # Detect format: if both separators present, identify which is decimal
    has_dot   = "." in s
    has_comma = "," in s

    if has_dot and has_comma:
        # Whichever appears last is the decimal separator
        last_dot   = s.rfind(".")
        last_comma = s.rfind(",")
        if last_comma > last_dot:
            # European format: 1.000,50 → remove dots, replace comma with dot
            s = s.replace(".", "").replace(",", ".")
        else:
            # Anglo format: 1,000.50 → remove commas
            s = s.replace(",", "")
    elif has_comma and not has_dot:
        # Could be decimal comma (1000,50) or thousands comma (1,500)
        # If there are exactly 3 digits after the comma, treat as thousands sep
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) == 3:
            s = s.replace(",", "")  # thousands separator
        else:
            s = s.replace(",", ".")  # decimal separator
    # If only dots, leave as-is

    try:
        return int(round(float(s) * 100))
    except (ValueError, TypeError):
        return None


def _parse_date(raw: str) -> datetime | None:
    formats = ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(raw.strip(), fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


async def _extract_transactions_with_llm(
    raw_text: str, settings
) -> list[dict[str, Any]]:
    """
    Use LLM (DeepSeek / OpenAI compatible) to extract structured transactions.
    Returns list of dicts: date, amount, type, description, vendor, confidence.
    """
    llm = ChatOpenAI(
        model=settings.llm_model,
        temperature=0.0,
        max_tokens=4096,
        api_key=settings.openai_api_key,
        base_url=settings.llm_base_url or None,
    )
    system = (
        "You are a financial data extraction specialist. "
        "Extract all financial transactions from the provided document text as a JSON array.\n"
        "For each transaction include:\n"
        "- date: transaction date (DD.MM.YYYY or YYYY-MM-DD)\n"
        "- amount: numeric amount (no currency symbol)\n"
        "- type: 'income' or 'expense'\n"
        "- description: transaction description\n"
        "- vendor: supplier or customer name (null if unknown)\n"
        "- confidence: confidence score 0-1 for this extraction\n\n"
        "Respond with ONLY a valid JSON array. No explanation, no markdown.\n"
        'Example: [{"date": "15.03.2024", "amount": 1500.00, "type": "expense", '
        '"description": "Electricity bill", "vendor": "City Power Co.", "confidence": 0.95}]'
    )
    messages = [
        SystemMessage(content=system),
        HumanMessage(content=f"Document text:\n\n{raw_text[:8000]}"),
    ]
    import json
    response = await llm.ainvoke(messages)
    content = response.content.strip()
    if content.startswith("```"):
        content = re.sub(r"^```[a-z]*\n?", "", content)
        content = re.sub(r"\n?```$", "", content)
    return json.loads(content)


def _read_pdf(file_path: str) -> str:
    """
    Read PDF using multi-strategy OCR pipeline.

    Tries native text extraction first (fast, high confidence).
    Falls back to pdfplumber for table-heavy documents.
    Falls back to Tesseract OCR for scanned/image-based PDFs.

    Returns the best-quality text available, with a warning comment
    prepended if confidence is low (for LLM fallback awareness).
    """
    try:
        from app.services.ocr_service import extract_text_from_pdf
        result = extract_text_from_pdf(file_path)

        if result.needs_llm_fallback:
            # Prepend low-confidence marker for LLM extraction path
            prefix = (
                f"[OCR_LOW_CONFIDENCE: {result.confidence:.0%} — "
                f"strategy={result.strategy_used}, pages={result.page_count}]\n\n"
            )
            return prefix + result.text

        # Append warnings for partial OCR
        text = result.text
        if result.warnings and result.confidence < 0.85:
            text += "\n\n[OCR_WARNINGS: " + "; ".join(result.warnings[:3]) + "]"

        return text

    except Exception as exc:
        # Graceful fallback to original simple extraction
        logger.warning("OCR service failed (%s), falling back to basic PDF read: %s", file_path, exc)
        import fitz
        text_parts: list[str] = []
        with fitz.open(file_path) as doc:
            for page in doc:
                text_parts.append(page.get_text())
        return "\n".join(text_parts)


def _read_excel(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
    return "\n".join(rows)


def _read_csv(file_path: str) -> str:
    with open(file_path, encoding="utf-8-sig", newline="") as f:
        return f.read()


def _statement_to_transactions(statement: ParsedStatement) -> list[dict[str, Any]]:
    """Convert ParsedStatement (bank parser output) to the CFO pipeline's transaction format."""
    transactions = []
    for tx in statement.transactions:
        transactions.append({
            "amount_cents": tx.amount_cents,
            "currency": tx.currency,
            "type": tx.tx_type,
            "category": _guess_category(tx.description),
            "description": tx.description,
            "vendor": tx.vendor,
            "transaction_date": tx.date.isoformat() if tx.date else None,
            "raw_text": tx.raw_row,
            "confidence": 0.95,  # structured parsers are high-confidence
        })
    return transactions


async def run_data_ingestion(
    state: CFOState, config: AgentRunConfig
) -> SkillResult:
    """
    Data Ingestion Skill.
    done_when: state['transactions'] is a non-empty list.

    Strategy:
      1. Read raw text from file (PDF / Excel / CSV)
      2. Try ParserRegistry (bank-specific rule-based parsers) — fast, free
      3. If no bank match → fall back to LLM extraction
    """
    settings = get_settings()
    file_path = state.get("file_path", "")
    file_type = state.get("file_type", "")

    if not os.path.exists(file_path):
        return SkillResult(ok=False, detail=f"File not found: {file_path}", halt=True)

    try:
        if file_type == "pdf":
            raw_text = _read_pdf(file_path)
        elif file_type in ("xlsx", "xls"):
            raw_text = _read_excel(file_path)
        elif file_type == "csv":
            raw_text = _read_csv(file_path)
        else:
            return SkillResult(ok=False, detail=f"Unsupported file type: {file_type}", halt=True)

        if not raw_text.strip():
            return SkillResult(
                ok=False,
                detail="Could not extract any text from the file.",
                needs_review=True,
                confidence=0.0,
            )

        # ── Strategy 1: Bank-specific rule-based parser ────────────────────
        detected_bank = ParserRegistry.detect(raw_text)
        if detected_bank is not None:
            logger.info(
                "job=%s — using bank parser: %s",
                state.get("job_id"), detected_bank.bank_display_name,
            )
            statement: ParsedStatement = detected_bank().parse(raw_text, file_path)
            transactions = _statement_to_transactions(statement)

            if transactions:
                parseable = sum(
                    1 for tx in transactions
                    if tx["amount_cents"] > 0 and tx["transaction_date"]
                )
                overall_confidence = 0.95 if parseable / len(transactions) >= 0.8 else 0.75
                detail = (
                    f"[{detected_bank.bank_display_name}] Parsed {len(transactions)} transactions "
                    f"({parseable} fully parsed, confidence={overall_confidence:.2f})"
                )
                if statement.parse_warnings:
                    logger.warning(
                        "job=%s — parser warnings: %s",
                        state.get("job_id"), statement.parse_warnings,
                    )
                return SkillResult(
                    ok=True,
                    patch={"raw_text": raw_text, "transactions": transactions},
                    confidence=overall_confidence,
                    needs_review=overall_confidence < 0.80,
                    detail=detail,
                )
            else:
                logger.warning(
                    "job=%s — bank parser returned 0 transactions, falling back to LLM",
                    state.get("job_id"),
                )

        # ── Strategy 2: LLM fallback ───────────────────────────────────────
        logger.info("job=%s — no bank parser matched, using LLM extraction", state.get("job_id"))
        raw_transactions = await _extract_transactions_with_llm(raw_text, settings)

        if not raw_transactions:
            return SkillResult(
                ok=False,
                detail="LLM returned no transactions from the document.",
                needs_review=True,
                confidence=0.4,
            )

        transactions = []
        confidences: list[float] = []

        for t in raw_transactions:
            amount_cents = _parse_amount(str(t.get("amount", "")))
            parsed_date = _parse_date(str(t.get("date", "")))
            conf = float(t.get("confidence", 0.8))
            confidences.append(conf)

            transactions.append({
                "amount_cents": amount_cents or 0,
                "currency": "USD",
                "type": t.get("type", "expense"),
                "category": _guess_category(t.get("description", "")),
                "description": t.get("description", ""),
                "vendor": t.get("vendor"),
                "transaction_date": parsed_date.isoformat() if parsed_date else None,
                "raw_text": str(t),
                "confidence": conf,
            })

        overall_confidence = min(confidences) if confidences else 0.5
        parseable = sum(
            1 for tx in transactions
            if tx["amount_cents"] > 0 and tx["transaction_date"]
        )
        parse_ratio = parseable / len(transactions) if transactions else 0

        if parse_ratio < 0.5:
            overall_confidence = min(overall_confidence, 0.55)
        elif parse_ratio < 0.8:
            overall_confidence = min(overall_confidence, 0.75)

        return SkillResult(
            ok=True,
            patch={"raw_text": raw_text, "transactions": transactions},
            confidence=overall_confidence,
            needs_review=overall_confidence < 0.80,
            detail=(
                f"[LLM] Extracted {len(transactions)} transactions "
                f"({parseable} fully parsed, confidence={overall_confidence:.2f})"
            ),
        )

    except Exception as exc:
        logger.exception("Data ingestion failed for job=%s", state.get("job_id"))
        return SkillResult(ok=False, detail=f"Ingestion error: {exc}", halt=True)
